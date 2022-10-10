from re import U
from openpyxl import load_workbook
import pyautogui as pg
import pyperclip
import random
import time
import wikipedia
import webbrowser
import os

# สั่งเปิด web
url = 'http://uncle-machine.com'
webbrowser.open(url)
time.sleep(2)
pg.hotkey('win','up',)# ขยายหน้าจอ หน้าต่าง
time.sleep(2)

# pg.press('tab', presses=7)
# userN = 'loong7777@gmail.com'
# passW =  '12345'
# pyperclip.copy(userN)
# pg.hotkey('ctrl', 'v')
# pg.press('tab')
# pyperclip.copy(passW)
# pg.hotkey('ctrl', 'v')
# pg.press('tab')
# pg.hotkey('enter')


path = r'C:\Users\bakpu\Desktop\Automate-bot'# ให้กดปุ่มที่ cut รูปมา
btn1 = os.path.join(path, 'product.png')
pg.click('product.png')
time.sleep(2)
# pg.click(r'C:\Users\bakpu\Desktop\Automate-bot\product.png')



pg.moveTo(100, 200)

excelfile = load_workbook('fruit.xlsx')
sheets = excelfile.active
# sheets2 = excelfile['sheets2']#การเลือก sheets
# sheets3 = excelfile[excelfile.sheetnames[0]]#การเลือก sheets

# print(sheets['B2'].value)

count = len(sheets['B'])
# print(count)

fruitlist = []

for i in range(2,count+1):
    data = sheets.cell(row=i, column=2).value
    # print(data)
    split = data.split(',')
    if len(split) >= 2:
        # print(split[0])
        if split[0] not in fruitlist:
            fruitlist.append(split[0]) 
            
    else:
        split = data.split('(')
        if split[0] not in fruitlist:
            # print(split)
            fruitlist.append(split[0])

# for f in fruitlist:
#     print(f)  

######### Work Flow ##########

#คลิกที่ เว็ปเบราเซอร์
pg.click(100, 200)

for f in fruitlist:

    #กด Tab 6 ครั้งเพื่อให้ เคอร์เซอร์อยู่ในช่องกรอก
    pg.press('tab', presses=6)

    #กรอกข้อมูลชื่อ + tab
    product = f
    pyperclip.copy(product)
    pg.hotkey('ctrl', 'v')
    pg.press('tab')

    #สุ่มราคา
    rand = random.choice(range(100, 1001, 100))

    #กรอกราคา (เป็นตัวเลข) + tab
    pyperclip.copy(rand)
    pg.hotkey('ctrl', 'v')
    pg.press('tab')

    #ข้อมูลสินค้า กรอกชื่อร้าน + tab
    try:
        info = wikipedia.summary(f)
    except:
        info = 'No Detail - Viboon Shop'
    pyperclip.copy(info)
    pg.hotkey('ctrl', 'v')
    pg.press('tab')

    # tab เพื่อไปยังปุ่ม Submit
    pg.press('tab')

    # กดปุ่ม  Enter เพื่อ submit
    pg.press('enter')
    time.sleep(1)
