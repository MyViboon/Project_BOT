from tkinter import Button
from xml.etree.ElementPath import xpath_tokenizer
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import wikipedia
import random
from openpyxl import load_workbook
import os

photopath = os.getcwd()
allflie = os.listdir(photopath)

############### DATA ####################
excelfile = load_workbook('fruit.xlsx')
sheets = excelfile.active
count = len(sheets['B'])
# print(count)

fruitlist = []

for i in range(2,count+1):
    data = sheets.cell(row=i, column=2).value
    split = data.split(',')
    if len(split) >= 2:
        if split[0] not in fruitlist:
            fruitlist.append(split[0].strip()) # .strip() คือสำหรับตัดช่องว่างชื่อไฟล์
    else:
        split = data.split('(')
        if split[0] not in fruitlist:
            fruitlist.append(split[0].strip()) # .strip() คือสำหรับตัดช่องว่างชื่อไฟล์

############################################

url = 'http://www.uncle-machine.com/login/'

path = r'C:\Users\bakpu\Desktop\Auto-Web\chromedriver_win32\chromedriver.exe'
driver  = webdriver.Chrome(path)
driver.get(url)

# search = driver.find_element("name","q")
# search.send_keys('ลำโพง')
# search.send_keys(Keys.RETURN) # การกดปุ่ม enter

########### LOGIN ##################
username = driver.find_element("id", "username")
username.send_keys('loong7777@gmail.com')

password = driver.find_element("id", "password")
password.send_keys('12345')

buttom = driver.find_element("xpath",'/html/body/div[2]/form/button')
buttom.click()

########### ADD PRODUCT#############

for f in fruitlist:
    url2 = "http://www.uncle-machine.com/addproduct/"
    driver.get(url2)

    name = driver.find_element("id", "name")
    name.send_keys(f)

    pc = random.choice([100, 300, 600, 500, 800])
    price = driver.find_element("id", "price")
    price.send_keys(pc)

    try:
        dt = wikipedia.summary(f)
    except:
        dt = 'ViboonFruit - Contact us for detail'

    detail = driver.find_element("id", "detail")
    detail.send_keys(dt)

    photo = driver.find_element("id", "photo")
    if f + '.jpg'in allflie: 
        photo_file = os.path.join(photopath, f+ '.jpg')
        photo.send_keys(photo_file)
    elif f + '.png' in allflie:
        photo_file = os.path.join(photopath, f+ '.png')
        photo.send_keys(photo_file)

    buttom = driver.find_element("xpath", "/html/body/div[2]/form/button")
    buttom.click()

# driver.close()




